import json
import re
from pathlib import Path
from openpyxl import load_workbook
from utils.number_utils import to_float, to_int

CATEGORIAS_DESPESAS = {"Viagens","Combustível","Hospedagem","Alimentação","Software","Plotagem","Impressões","Veículos","Passagens","Taxas","Cartório","Outros"}

EXPECTED_HEADERS = {
    "Orcamento_Inicial": ["Contrato_Codigo","Competencia","Versao","Valor_Receita","Centro_Custo","Observacoes"],
    "Revisao_Orcamento": ["Contrato_Codigo","Competencia","Versao","Valor_Receita","Centro_Custo","Observacoes"],
    "Faturamento_Mensal": ["Contrato_Codigo","Competencia","Documento_Ref","Valor_Faturado","Impostos_%","Observacoes"],
    "Despesas_Previstas": ["Contrato_Codigo","Competencia","Categoria","Descricao","Fornecedor","Valor_Previsto","Centro_Custo","Observacoes"],
    "Despesas_Realizadas": ["Contrato_Codigo","Competencia","Categoria","Descricao","Fornecedor","Valor_Realizado","Documento_Ref","Centro_Custo","Observacoes"],
}

COMP_RX = re.compile(r"^(0[1-9]|1[0-2])/\d{4}$")

class FinanceImportService:
    def __init__(self, import_repo, despesas_plan_service, despesas_real_service):
        self.import_repo = import_repo
        self.despesas_plan_service = despesas_plan_service
        self.despesas_real_service = despesas_real_service

    def _check_headers(self, ws, expected):
        headers = [str(c.value or "").strip() for c in ws[1][:len(expected)]]
        return headers == expected, headers

    def _contract_id(self, codigo):
        row = self.import_repo.get_contrato_by_codigo(str(codigo).strip())
        return row["id"] if row else None

    def validate_workbook(self, filepath):
        wb = load_workbook(filepath, data_only=True)
        errors = []
        counts = {k: 0 for k in EXPECTED_HEADERS.keys()}

        for sheet_name, headers in EXPECTED_HEADERS.items():
            if sheet_name not in wb.sheetnames:
                errors.append(f"Aba obrigatória ausente: {sheet_name}.")
                continue
            ws = wb[sheet_name]
            ok, found = self._check_headers(ws, headers)
            if not ok:
                errors.append(f"Aba {sheet_name}: cabeçalhos inválidos. Esperado: {headers}. Encontrado: {found}.")

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row is None or all(v in (None, "") for v in row):
                    continue
                counts[sheet_name] += 1
                codigo = str(row[0] or "").strip()
                comp = str(row[1] or "").strip()
                if not codigo:
                    errors.append(f"{sheet_name} linha {row_idx}: informar Contrato_Codigo.")
                    continue
                if not self._contract_id(codigo):
                    errors.append(f"{sheet_name} linha {row_idx}: contrato '{codigo}' não encontrado.")
                if not COMP_RX.match(comp):
                    errors.append(f"{sheet_name} linha {row_idx}: competência inválida '{comp}'. Use MM/AAAA.")

                if sheet_name in {"Orcamento_Inicial", "Revisao_Orcamento"}:
                    if to_int(row[2]) <= 0:
                        errors.append(f"{sheet_name} linha {row_idx}: Versao deve ser inteiro > 0.")
                    if to_float(row[3]) < 0:
                        errors.append(f"{sheet_name} linha {row_idx}: Valor_Receita não pode ser negativo.")
                elif sheet_name == "Faturamento_Mensal":
                    if to_float(row[3]) < 0:
                        errors.append(f"{sheet_name} linha {row_idx}: Valor_Faturado não pode ser negativo.")
                elif sheet_name in {"Despesas_Previstas", "Despesas_Realizadas"}:
                    categoria = str(row[2] or "").strip()
                    if categoria not in CATEGORIAS_DESPESAS:
                        errors.append(f"{sheet_name} linha {row_idx}: categoria inválida '{categoria}'.")
                    if sheet_name == "Despesas_Previstas":
                        if to_float(row[5]) < 0:
                            errors.append(f"{sheet_name} linha {row_idx}: Valor_Previsto não pode ser negativo.")
                    else:
                        if to_float(row[5]) < 0:
                            errors.append(f"{sheet_name} linha {row_idx}: Valor_Realizado não pode ser negativo.")

        return {
            "ok": len(errors) == 0,
            "errors": errors,
            "counts": counts,
            "summary": f"{sum(counts.values())} linhas lidas; {len(errors)} erro(s) encontrado(s).",
        }

    def import_workbook(self, filepath):
        report = self.validate_workbook(filepath)
        name = Path(filepath).name
        if not report["ok"]:
            self.import_repo.create_log(name, "erro_validacao", report["summary"], json.dumps(report, ensure_ascii=False))
            return report

        wb = load_workbook(filepath, data_only=True)
        inserted = {"orcamento": 0, "faturamento": 0, "despesas_previstas": 0, "despesas_realizadas": 0}

        for sheet_name in ["Orcamento_Inicial", "Revisao_Orcamento"]:
            ws = wb[sheet_name]
            tipo = "inicial" if sheet_name == "Orcamento_Inicial" else "revisado"
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row is None or all(v in (None, "") for v in row):
                    continue
                contrato_id = self._contract_id(row[0])
                self.import_repo.upsert_orcamento({
                    "contrato_id": contrato_id,
                    "competencia": str(row[1]).strip(),
                    "versao": to_int(row[2]),
                    "tipo_orcamento": tipo,
                    "valor_receita": to_float(row[3]),
                    "centro_custo": str(row[4] or "").strip(),
                    "observacoes": str(row[5] or "").strip(),
                    "fonte": "importacao_planilha",
                })
                inserted["orcamento"] += 1

        ws = wb["Faturamento_Mensal"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None or all(v in (None, "") for v in row):
                continue
            contrato_id = self._contract_id(row[0])
            self.import_repo.upsert_faturamento({
                "contrato_id": contrato_id,
                "competencia": str(row[1]).strip(),
                "documento_ref": str(row[2] or "").strip(),
                "valor_faturado": to_float(row[3]),
                "impostos_percentual": to_float(row[4] or 16.8),
                "observacoes": str(row[5] or "").strip(),
                "fonte": "importacao_planilha",
            })
            inserted["faturamento"] += 1

        ws = wb["Despesas_Previstas"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None or all(v in (None, "") for v in row):
                continue
            contrato_id = self._contract_id(row[0])
            self.despesas_plan_service.repo.upsert_prevista_import({
                "contrato_id": contrato_id,
                "competencia": str(row[1]).strip(),
                "categoria": str(row[2] or "").strip(),
                "descricao": str(row[3] or "").strip(),
                "fornecedor": str(row[4] or "").strip(),
                "valor_previsto": to_float(row[5]),
                "centro_custo": str(row[6] or "").strip(),
                "observacoes": str(row[7] or "").strip(),
            })
            inserted["despesas_previstas"] += 1

        ws = wb["Despesas_Realizadas"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None or all(v in (None, "") for v in row):
                continue
            contrato_id = self._contract_id(row[0])
            self.despesas_real_service.repo.upsert_realizada_import({
                "contrato_id": contrato_id,
                "competencia": str(row[1]).strip(),
                "categoria": str(row[2] or "").strip(),
                "descricao": str(row[3] or "").strip(),
                "fornecedor": str(row[4] or "").strip(),
                "valor_realizado": to_float(row[5]),
                "documento_ref": str(row[6] or "").strip(),
                "centro_custo": str(row[7] or "").strip(),
                "observacoes": str(row[8] or "").strip(),
            })
            inserted["despesas_realizadas"] += 1

        result = {
            "ok": True,
            "errors": [],
            "counts": report["counts"],
            "inserted": inserted,
            "summary": (
                f"Importação concluída: orçamento={inserted['orcamento']}, "
                f"faturamento={inserted['faturamento']}, "
                f"despesas previstas={inserted['despesas_previstas']}, "
                f"despesas realizadas={inserted['despesas_realizadas']}."
            ),
        }
        self.import_repo.create_log(name, "sucesso", result["summary"], json.dumps(result, ensure_ascii=False))
        return result

    def recent_logs(self):
        return self.import_repo.list_logs()
