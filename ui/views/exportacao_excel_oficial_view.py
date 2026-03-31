from PySide6.QtWidgets import QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QFileDialog, QMessageBox, QTableWidget
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from ui.widgets.common import configure_table
from ui.widgets.table_items import TextItem

class ExportacaoExcelOficialView(QWidget):
    def __init__(self, service):
        super().__init__()
        self.service = service
        self.payload = {}
        self._build_ui()
        self.reload_data()

    def _build_ui(self):
        root = QVBoxLayout(self)
        title = QLabel("Exportação Excel Oficial")
        title.setStyleSheet("font-size:18px;font-weight:700;")
        root.addWidget(title)
        bar = QHBoxLayout()
        self.btn_atualizar = QPushButton("Atualizar")
        self.btn_atualizar.clicked.connect(self.reload_data)
        self.btn_exportar = QPushButton("Exportar Excel Oficial")
        self.btn_exportar.clicked.connect(self.exportar_excel)
        bar.addWidget(self.btn_atualizar)
        bar.addWidget(self.btn_exportar)
        bar.addStretch()
        root.addLayout(bar)
        self.table = QTableWidget(0, 2)
        self.table.setHorizontalHeaderLabels(["Aba", "Quantidade de linhas"])
        configure_table(self.table)
        root.addWidget(self.table)

    def reload_data(self):
        self.payload = self.service.pacote_oficial_dict()
        self.table.setRowCount(0)
        for aba, rows in self.payload.items():
            r = self.table.rowCount()
            self.table.insertRow(r)
            self.table.setItem(r, 0, TextItem(aba))
            self.table.setItem(r, 1, TextItem(str(len(rows))))

    def exportar_excel(self):
        path, _ = QFileDialog.getSaveFileName(self, "Salvar Excel Oficial", "pacote_financeiro_oficial.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        try:
            wb = Workbook()
            first = True
            for aba, rows in self.payload.items():
                ws = wb.active if first else wb.create_sheet(title=aba[:31])
                first = False
                ws.title = aba[:31]
                if rows:
                    headers = list(rows[0].keys())
                    ws.append(headers)
                    for row in rows:
                        ws.append([row.get(h, "") for h in headers])
                    for cell in ws[1]:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill("solid", fgColor="D9EAF7")
                else:
                    ws.append(["Sem dados"])
            wb.save(path)
            QMessageBox.information(self, "Sucesso", f"Excel oficial exportado para:\n{path}")
        except Exception as exc:
            QMessageBox.warning(self, "Erro", str(exc))
