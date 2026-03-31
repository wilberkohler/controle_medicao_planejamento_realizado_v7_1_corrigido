from pathlib import Path
from PySide6.QtWidgets import QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QFileDialog, QMessageBox, QTableWidget
from ui.widgets.common import configure_table
from ui.widgets.table_items import TextItem
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

class ExportacaoExcelView(QWidget):
    def __init__(self, service):
        super().__init__()
        self.service = service
        self.payload = {}
        self._build_ui()
        self.reload_data()

    def _build_ui(self):
        root = QVBoxLayout(self)
        title = QLabel("Exportação Excel Consolidada")
        title.setStyleSheet("font-size:18px;font-weight:700;")
        root.addWidget(title)

        info = QLabel("Gera um único arquivo Excel com múltiplas abas para contador, diretoria e planejamento anual.")
        info.setWordWrap(True)
        root.addWidget(info)

        bar = QHBoxLayout()
        self.btn_atualizar = QPushButton("Atualizar")
        self.btn_atualizar.clicked.connect(self.reload_data)
        self.btn_exportar = QPushButton("Exportar Excel")
        self.btn_exportar.clicked.connect(self.exportar_excel)
        bar.addWidget(self.btn_atualizar)
        bar.addWidget(self.btn_exportar)
        bar.addStretch()
        root.addLayout(bar)

        self.table = QTableWidget(0, 2)
        self.table.setHorizontalHeaderLabels(["Aba", "Quantidade de linhas"])
        configure_table(self.table)
        root.addWidget(self.table)

    def reload_data(self):
        self.payload = self.service.pacote_anual_dict()
        self.table.setRowCount(0)
        for aba, rows in self.payload.items():
            r = self.table.rowCount()
            self.table.insertRow(r)
            self.table.setItem(r, 0, TextItem(aba))
            self.table.setItem(r, 1, TextItem(str(len(rows))))

    def exportar_excel(self):
        path, _ = QFileDialog.getSaveFileName(self, "Salvar Excel", "pacote_contabil_anual.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        try:
            wb = Workbook()
            first = True
            for aba, rows in self.payload.items():
                ws = wb.active if first else wb.create_sheet(title=aba[:31])
                first = False
                ws.title = aba[:31]

                if rows:
                    headers = list(rows[0].keys())
                    ws.append(headers)
                    for row in rows:
                        ws.append([row.get(h, "") for h in headers])

                    for cell in ws[1]:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill("solid", fgColor="D9EAF7")

                    for col in ws.columns:
                        max_len = 0
                        col_letter = col[0].column_letter
                        for cell in col:
                            max_len = max(max_len, len(str(cell.value or "")))
                        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 28)
                else:
                    ws.append(["Sem dados"])

            if "README" not in wb.sheetnames:
                ws = wb.create_sheet(title="README")
                ws["A1"] = "Pacote contábil anual"
                ws["A2"] = "Abas:"
                row = 3
                for aba in self.payload.keys():
                    ws[f"A{row}"] = aba
                    row += 1

            wb.save(path)
            QMessageBox.information(self, "Sucesso", f"Excel exportado para:\n{path}")
        except Exception as exc:
            QMessageBox.warning(self, "Erro", str(exc))
